import openpyxl


def get_sheet(location:str,sheetname:str):
    wk=openpyxl.load_workbook(location)
    sh=wk[sheetname]
    return sh



def fetch_rows_from_excel(location:str,sheetname:str)->int:
    sh=get_sheet(location,sheetname)
    max_row=sh.max_row
    return max_row

def fetch_all_row_column(location:str, sheetname:str, row:int,column:int)->str:
    sh=get_sheet(location,sheetname)
    return sh.cell(row=int(row),column=int(column)).value
    


# if __name__ == '__main__':
#     print(fetch_all_row_column(location="C:\\Users\\psara\Downloads\\Python_Robot_Framework_Testing\\External_python\\book3.xlsx",sheetname="Sheet1",row=1,column=1))
    