import openpyxl
from functools import reduce

wk = openpyxl.Workbook()
sh = wk.active

value_list = [["pratik", "pratik123", "psarangi50@fmail.com"], ["abi", "rika1997", "rika1997@gmail.com"]]

s = lambda: [x for item in value_list for x in item]
print(s())
count = 0

for pos, val in enumerate(s()):
    if pos < 3:
        sh.cell(row=1, column=pos + 1).value = val
    else:
        sh.cell(row=2, column=count + 1).value = val
        count += 1

wk.save("book3.xlsx")
